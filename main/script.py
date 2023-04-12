'''      code done by Fips5, 
        https://github.com/Fips5        '''

import os
from openpyxl import load_workbook

while True:
    dir_path = "./files(example folder)"  # CHANGE THE PATH FOR THE FOLDER YOU HAVE, PUT THE SPECIFIC PATH

    xlsx_files = [f for f in os.listdir(dir_path) if f.endswith(".xlsx")]

    choice = input("Do you want to check a row (r), column (c), or specific cell (s)? ")
    if choice.lower() == "r":
        row_num = int(input("Enter the row number: "))
        col_letter = None
    elif choice.lower() == "c":
        col_letter = input("Enter the column letter: ")
        row_num = None
    elif choice.lower() == "s":
        col_letter = input("Enter the column letter: ")
        row_num = int(input("Enter the row number: "))
    else:
        print("Invalid choice. Please enter r, c, or s.")
        continue

    cell_values = {}

    for filename in xlsx_files:
        filepath = os.path.join(dir_path, filename)
        workbook = load_workbook(filepath)
        sheet = workbook.active
        if row_num is not None:
            values = []
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for c in cell:
                    values.append(c.value)
            cell_values[filename] = values
        elif col_letter is not None:
            values = []
            col_index = ord(col_letter) - 96
            for cell in sheet.iter_cols(min_col=col_index, max_col=col_index):
                for c in cell:
                    values.append(c.value)
            cell_values[filename] = values
        else:
            cell = sheet["{}{}".format(col_letter, row_num)]
            cell_values[filename] = [cell.value]

    all_values_same = all(value == list(cell_values.values())[0] for value in cell_values.values())

    if all_values_same:
        print("All files have the same value: {}".format(list(cell_values.values())[0][0]))
    else:
        for filename, values in cell_values.items():
            if len(set(values)) == 1:
                print("{}: All values are the same: {}".format(filename, values[0]))
            else:
                for idx, val in enumerate(values):
                    if val != values[0]:
                        print("{}: Cell {}{} has a different value: {}".format(filename, col_letter, row_num+idx, val))

        normal_value = list(cell_values.values())[0][0]
        print("Normal value from other files: {}".format(normal_value))

    user_choice = input("Do you want to exit? y/n: ")
    if user_choice.lower() == "y":
        break
    elif user_choice.lower() == "n":
        continue
    else:
        print("Invalid choice. Please enter y or n.")

print("Program has ended.")

